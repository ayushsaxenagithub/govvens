"""
Enhanced website views with comprehensive error handling, analytics, and security features
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Count, Avg, Max, Min, Sum, F
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from .models import Event, Booking
from user.models import User, NotificationSubscription, UserSession, UserActivity
from django.utils.dateparse import parse_date
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import logging

logger = logging.getLogger(__name__)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def is_superuser(user):
    """Check if user is authenticated superuser"""
    return user.is_authenticated and user.is_superuser

def safe_get_stats(model_class, filters=None):
    """Safely get statistics from a queryset"""
    try:
        if filters:
            qs = model_class.objects.filter(**filters)
        else:
            qs = model_class.objects.all()
        return {
            'total': qs.count(),
            'exists': qs.exists()
        }
    except Exception as e:
        logger.error(f"Error getting stats for {model_class.__name__}: {str(e)}")
        return {'total': 0, 'exists': False}

# ============================================================================
# PUBLIC PAGES
# ============================================================================

def landing_page(request):
    """Landing page with real event data"""
    try:
        upcoming_events = Event.objects.filter(is_active=True).order_by('date', 'time')[:3]
        
        featured_events = []
        for event in upcoming_events:
            featured_events.append({
                'id': event.id,
                'name': event.name,
                'date': event.date.strftime('%d %B %Y'),
                'time': event.time.strftime('%I:%M %p'),
                'stadium': event.stadium,
                'ticket_price': f'₹{int(event.ticket_price)}',
                'available_seats': event.available_seats
            })
        
        return render(request, 'website/landing.html', {
            'page_title': 'GOVVENS - Crowd-Safe Ticketing',
            'featured_events': featured_events,
            'has_events': len(featured_events) > 0
        })
    except Exception as e:
        logger.error(f"Error in landing_page: {str(e)}")
        return render(request, 'website/landing.html', {
            'page_title': 'GOVVENS - Crowd-Safe Ticketing',
            'featured_events': [],
            'has_events': False,
            'error': 'Unable to load featured events'
        })

def events_list(request):
    """Events list page with pagination"""
    try:
        events = Event.objects.filter(is_active=True).order_by('date', 'time')
        
        if not events.exists():
            events_data = [
                {
                    'id': 1,
                    'name': 'India vs Australia',
                    'date': '2025-11-05',
                    'time': '18:00',
                    'stadium': 'Chinnaswamy Stadium',
                    'ticket_price': '₹2500',
                    'available_seats': '1200'
                },
                {
                    'id': 2,
                    'name': 'IPL Final',
                    'date': '2025-11-10',
                    'time': '19:00',
                    'stadium': 'Chinnaswamy Stadium',
                    'ticket_price': '₹5000',
                    'available_seats': '800'
                },
            ]
        else:
            events_data = [{
                'id': e.id,
                'name': e.name,
                'date': e.date.strftime('%Y-%m-%d'),
                'time': e.time.strftime('%H:%M'),
                'stadium': e.stadium,
                'ticket_price': f'₹{e.ticket_price}',
                'available_seats': str(e.available_seats)
            } for e in events]
        
        return render(request, 'website/events_list.html', {
            'page_title': 'All Upcoming Matches',
            'events': events_data
        })
    except Exception as e:
        logger.error(f"Error in events_list: {str(e)}")
        messages.error(request, 'Error loading events')
        return render(request, 'website/events_list.html', {
            'page_title': 'All Upcoming Matches',
            'events': [],
            'error': str(e)
        })

def event_detail(request, event_id):
    """Event details page"""
    try:
        event = Event.objects.get(id=event_id, is_active=True)
        event_data = {
            'id': event.id,
            'name': event.name,
            'date': event.date.strftime('%Y-%m-%d'),
            'time': event.time.strftime('%H:%M'),
            'stadium': event.stadium,
            'ticket_price': f'₹{event.ticket_price}',
            'available_seats': str(event.available_seats),
        }
        
        subscription_status = False
        if request.user.is_authenticated:
            subscription_status = NotificationSubscription.objects.filter(
                user=request.user, event_id=event_id
            ).exists()
        
        return render(request, 'website/events_detail.html', {
            'page_title': f'{event_data["name"]} - Match Details',
            'event': event_data,
            'subscription_status': subscription_status
        })
    except Event.DoesNotExist:
        logger.warning(f"Event {event_id} not found")
        messages.error(request, 'Event not found')
        return redirect('events_list')
    except Exception as e:
        logger.error(f"Error in event_detail: {str(e)}")
        messages.error(request, 'Error loading event details')
        return redirect('events_list')

def seat_selection(request, event_id):
    """Seat selection page"""
    try:
        if request.user.is_authenticated and not request.user.is_verified:
            messages.info(request, 'You can browse seats, but please verify your identity to complete booking.')
        
        try:
            event = Event.objects.get(id=event_id, is_active=True)
            event_data = {
                'id': event.id,
                'name': event.name,
                'date': event.date.strftime('%Y-%m-%d'),
                'time': event.time.strftime('%H:%M'),
                'stadium': event.stadium,
                'ticket_price': float(event.ticket_price)
            }
        except Event.DoesNotExist:
            messages.error(request, 'Event not found')
            return redirect('events_list')
        
        seat_map = {
            'blocks': [
                {'name': 'A', 'rows': 10, 'seats_per_row': 20},
                {'name': 'B', 'rows': 8, 'seats_per_row': 25},
                {'name': 'C', 'rows': 6, 'seats_per_row': 30}
            ]
        }
        
        if request.method == 'POST':
            if not request.user.is_authenticated:
                messages.warning(request, 'Please login to proceed with booking.')
                request.session['redirect_after_login'] = request.path
                return redirect('signup_login')
            if not request.user.is_verified:
                messages.warning(request, 'Please verify your identity to complete booking.')
                return redirect('verify_identity')
            
            selected_seats = request.POST.getlist('selected_seats')
            if selected_seats:
                request.session['selected_seats'] = selected_seats
                request.session['event_id'] = event_id
                return redirect('payment', event_id=event_id)
            else:
                messages.error(request, 'Please select at least one seat.')
        
        return render(request, 'website/seat_selection.html', {
            'page_title': f'Select Your Seats for {event_data["name"]}',
            'event': event_data,
            'seat_map': seat_map
        })
    except Exception as e:
        logger.error(f"Error in seat_selection: {str(e)}")
        messages.error(request, 'Error in seat selection')
        return redirect('events_list')

def payment(request, event_id):
    """Payment page"""
    try:
        demo_selected_seats = ['A-5-12', 'A-5-13', 'A-5-14']
        
        if not request.user.is_authenticated:
            try:
                event = Event.objects.get(id=event_id, is_active=True)
                ticket_price = float(event.ticket_price)
                event_data = {
                    'id': event.id,
                    'name': event.name,
                    'date': event.date.strftime('%Y-%m-%d'),
                    'time': event.time.strftime('%H:%M'),
                    'stadium': event.stadium,
                    'ticket_price': ticket_price
                }
            except Event.DoesNotExist:
                ticket_price = 2500.0
                event_data = {
                    'id': event_id,
                    'name': 'Demo Event',
                    'date': '2025-11-05',
                    'time': '18:00',
                    'stadium': 'Demo Stadium',
                    'ticket_price': ticket_price
                }
            
            total_price = len(demo_selected_seats) * ticket_price
            return render(request, 'website/payment.html', {
                'page_title': f'Review & Pay (Demo)',
                'event': event_data,
                'selected_seats': demo_selected_seats,
                'total_price': total_price,
                'is_demo': True
            })
        
        if not request.user.is_verified:
            messages.warning(request, 'Please verify your identity to complete booking.')
            return redirect('verify_identity')
        
        selected_seats = request.session.get('selected_seats', [])
        if not selected_seats:
            messages.error(request, 'No seats selected.')
            return redirect('seat_selection', event_id=event_id)
        
        try:
            event = Event.objects.get(id=event_id, is_active=True)
            ticket_price = float(event.ticket_price)
            event_data = {
                'id': event.id,
                'name': event.name,
                'date': event.date.strftime('%Y-%m-%d'),
                'time': event.time.strftime('%H:%M'),
                'stadium': event.stadium,
                'ticket_price': ticket_price
            }
        except Event.DoesNotExist:
            messages.error(request, 'Event not found')
            return redirect('events_list')
        
        total_price = len(selected_seats) * ticket_price
        
        if request.method == 'POST':
            payment_method = request.POST.get('payment_method')
            if payment_method:
                messages.success(request, 'Payment successful!')
                return redirect('ticket_confirmation')
            else:
                messages.error(request, 'Please select a payment method.')
        
        return render(request, 'website/payment.html', {
            'page_title': f'Review & Pay',
            'event': event_data,
            'selected_seats': selected_seats,
            'total_price': total_price,
            'is_demo': False
        })
    except Exception as e:
        logger.error(f"Error in payment: {str(e)}")
        messages.error(request, 'Error processing payment')
        return redirect('events_list')

def ticket_confirmation(request):
    """Ticket confirmation page"""
    try:
        if not request.user.is_authenticated:
            booking_data = {
                'id': 1,
                'event_name': 'India vs Australia',
                'date': '2025-11-05',
                'time': '18:00',
                'stadium': 'Chinnaswamy Stadium',
            }
            return render(request, 'website/ticket_confirmation.html', {
                'page_title': 'Ticket Confirmed! (Demo)',
                'booking': booking_data,
                'is_demo': True
            })
        
        selected_seats = request.session.get('selected_seats', [])
        event_id = request.session.get('event_id')
        
        if not selected_seats or not event_id:
            messages.error(request, 'No booking found.')
            return redirect('events_list')
        
        try:
            event = Event.objects.get(id=event_id)
            booking_data = {
                'id': 1,
                'event_name': event.name,
                'date': event.date.strftime('%Y-%m-%d'),
                'time': event.time.strftime('%H:%M'),
                'stadium': event.stadium,
            }
        except Event.DoesNotExist:
            booking_data = {
                'id': 1,
                'event_name': 'Event',
                'date': '2025-11-05',
                'time': '18:00',
                'stadium': 'Stadium',
            }
        
        if 'selected_seats' in request.session:
            del request.session['selected_seats']
        if 'event_id' in request.session:
            del request.session['event_id']
        
        return render(request, 'website/ticket_confirmation.html', {
            'page_title': 'Ticket Confirmed!',
            'booking': booking_data,
            'is_demo': False
        })
    except Exception as e:
        logger.error(f"Error in ticket_confirmation: {str(e)}")
        return redirect('events_list')

def my_tickets(request):
    """My tickets page"""
    try:
        if not request.user.is_authenticated:
            demo_bookings = [
                {
                    'id': 1,
                    'event_name': 'India vs Australia',
                    'date': '2025-11-05',
                    'time': '18:00',
                },
            ]
            return render(request, 'website/my_tickets.html', {
                'page_title': 'My Booked Tickets (Demo)',
                'bookings': demo_bookings,
                'is_demo': True
            })
        
        bookings = Booking.objects.filter(user=request.user).order_by('-id')[:50]
        
        bookings_data = []
        if bookings.exists():
            for b in bookings:
                try:
                    bookings_data.append({
                        'id': b.id,
                        'event_name': b.event.name if b.event else 'Unknown Event',
                        'date': b.event.date.strftime('%Y-%m-%d') if b.event else 'Unknown',
                        'time': b.event.time.strftime('%H:%M') if b.event else 'Unknown',
                    })
                except Exception as be:
                    logger.warning(f"Error processing booking {b.id}: {str(be)}")
                    continue
        
        return render(request, 'website/my_tickets.html', {
            'page_title': 'My Booked Tickets',
            'bookings': bookings_data if bookings_data else [],
            'is_demo': False
        })
    except Exception as e:
        logger.error(f"Error in my_tickets: {str(e)}")
        messages.error(request, 'Error loading tickets')
        return render(request, 'website/my_tickets.html', {
            'page_title': 'My Booked Tickets',
            'bookings': [],
            'error': str(e)
        })

def entry_exit_plan(request):
    """Map and entry/exit plan page"""
    try:
        user_pin = request.user.pin_code if request.user.is_authenticated else '560001'
        return render(request, 'website/map_entry_exit.html', {
            'page_title': 'Plan Your Journey to the Stadium',
            'user_pin': user_pin,
        })
    except Exception as e:
        logger.error(f"Error in entry_exit_plan: {str(e)}")
        return render(request, 'website/map_entry_exit.html', {
            'page_title': 'Plan Your Journey to the Stadium',
            'error': str(e)
        })

def exit_info(request):
    """Exit info page"""
    try:
        return render(request, 'website/exit_info.html', {
            'page_title': 'Exit Instructions',
        })
    except Exception as e:
        logger.error(f"Error in exit_info: {str(e)}")
        messages.error(request, 'Error loading exit information')
        return render(request, 'website/exit_info.html', {
            'page_title': 'Exit Instructions',
            'error': str(e)
        })

def faq_page(request):
    """FAQ page"""
    try:
        faqs = [
            {
                'question': 'What is Govvens?',
                'answer': 'Govvens is a civic-tech ticketing and crowd-safety platform that manages crowd entry, ticketing, and exit for large events using staggered timings and verified ticketing. We ensure safe, organized, and fair access to high-footfall events.',
                'category': 'general',
                'icon': 'bi-info-circle'
            },
            {
                'question': 'How does the notification system work?',
                'answer': "Users pay a small, refundable fee (₹10) to get SMS alerts when tickets for specific matches open. This ensures fair and early access. The fee is deductible from your ticket price when you make a booking. You'll receive timely notifications via SMS and email before tickets go on sale.",
                'category': 'notifications',
                'icon': 'bi-bell'
            },
            {
                'question': 'How are tickets verified?',
                'answer': "Each ticket is tied to the user through Aadhaar-based eKYC and facial recognition. At the venue, you'll go through a quick face/eye scan similar to DigiYatra. Backup verification uses mobile OTP and ID at entry gates if needed. This ensures secure entry and prevents ticket fraud.",
                'category': 'verification',
                'icon': 'bi-shield-check'
            },
            {
                'question': 'How is crowd control managed?',
                'answer': "The system assigns specific time slots for entry and exit through automated SMS and app notifications, minimizing congestion and stampede risk. Each user gets a personalized entry window (e.g., 17:00-17:30) and exit window, ensuring staggered flow of crowds throughout the event.",
                'category': 'crowd-control',
                'icon': 'bi-people'
            },
            {
                'question': 'Do I need to share my live location?',
                'answer': "No. Only your home PIN code is required. Routes and timings are computed using Google Maps API based on that. We calculate the best departure time from your location to arrive during your assigned entry window. Your privacy is protected - we never track your live location.",
                'category': 'privacy',
                'icon': 'bi-geo-alt'
            },
            {
                'question': 'Can I book seats for my friends or family?',
                'answer': "Yes. The seat map supports individual, friends, and family bookings. You can select multiple seats together, and they will be held for up to 15 minutes during checkout. Each person will need their own identity verification for entry at the venue.",
                'category': 'booking',
                'icon': 'bi-people-fill'
            },
            {
                'question': 'How do I pay for tickets?',
                'answer': "Payments are processed securely via Razorpay. We accept Credit Cards, Debit Cards, UPI, Net Banking, and Wallets. All transactions are encrypted and secure. You'll receive a payment confirmation and digital ticket with QR code immediately after successful payment.",
                'category': 'payment',
                'icon': 'bi-credit-card'
            },
            {
                'question': 'What if my payment fails?',
                'answer': "The reserved seats are released after 15 minutes if payment isn't completed. You'll receive a notification if your payment fails. You can retry the payment within the 15-minute window. If the time expires, you'll need to select seats again. We recommend having a backup payment method ready.",
                'category': 'payment',
                'icon': 'bi-exclamation-triangle'
            },
        ]
        return render(request, 'website/faq.html', {
            'page_title': 'Frequently Asked Questions',
            'faqs': faqs
        })
    except Exception as e:
        logger.error(f"Error in faq_page: {str(e)}")
        return render(request, 'website/faq.html', {
            'page_title': 'Frequently Asked Questions',
            'faqs': [],
            'error': str(e)
        })

# ============================================================================
# ADMIN DASHBOARDS
# ============================================================================

@login_required
@user_passes_test(is_superuser, login_url='/')
def dashboard(request):
    """Admin dashboard for managing users"""
    try:
        users = User.objects.all().order_by('-date_joined')
        
        search_query = request.GET.get('search', '')
        if search_query:
            users = users.filter(
                Q(email__icontains=search_query) |
                Q(username__icontains=search_query) |
                Q(phone__icontains=search_query)
            )
        
        filter_type = request.GET.get('filter', '')
        if filter_type == 'verified':
            users = users.filter(is_verified=True)
        elif filter_type == 'unverified':
            users = users.filter(is_verified=False)
        elif filter_type == 'active':
            users = users.filter(is_active=True)
        elif filter_type == 'inactive':
            users = users.filter(is_active=False)
        
        sort_by = request.GET.get('sort', '-date_joined')
        if sort_by in ['email', 'username', 'date_joined', '-date_joined']:
            users = users.order_by(sort_by)
        
        paginator = Paginator(users, 25)
        page_number = request.GET.get('page', 1)
        try:
            page_obj = paginator.page(page_number)
        except (PageNotAnInteger, EmptyPage):
            page_obj = paginator.page(1)
        
        # Statistics with error handling
        try:
            total_users = User.objects.count()
            verified_users = User.objects.filter(is_verified=True).count()
            active_users = User.objects.filter(is_active=True).count()
        except Exception as e:
            logger.error(f"Error getting user stats: {str(e)}")
            total_users = verified_users = active_users = 0
        
        context = {
            'page_title': 'Admin Dashboard - User Management',
            'users': page_obj,
            'total_users': total_users,
            'verified_users': verified_users,
            'active_users': active_users,
            'search_query': search_query,
            'filter_type': filter_type,
            'sort_by': sort_by,
        }
        
        return render(request, 'website/dashboard.html', context)
    except Exception as e:
        logger.error(f"Error in dashboard: {str(e)}")
        messages.error(request, 'Error loading dashboard')
        return render(request, 'website/dashboard.html', {
            'page_title': 'Admin Dashboard',
            'error': str(e),
            'users': [],
        })

@login_required
@user_passes_test(is_superuser, login_url='/')
@require_http_methods(["POST"])
def delete_user(request, user_id):
    """Delete a user"""
    try:
        user = get_object_or_404(User, id=user_id)
        if user.id == request.user.id:
            messages.error(request, 'You cannot delete your own account!')
        else:
            user_email = user.email
            user.delete()
            messages.success(request, f'User {user_email} deleted successfully.')
    except Exception as e:
        logger.error(f"Error deleting user: {str(e)}")
        messages.error(request, f'Error: {str(e)}')
    return redirect('dashboard')

@login_required
@user_passes_test(is_superuser, login_url='/')
@require_http_methods(["POST"])
def toggle_user_status(request, user_id):
    """Toggle user active status"""
    try:
        user = get_object_or_404(User, id=user_id)
        if user.id == request.user.id:
            messages.error(request, 'You cannot deactivate your own account!')
        else:
            user.is_active = not user.is_active
            user.save()
            status = 'activated' if user.is_active else 'deactivated'
            messages.success(request, f'User {user.email} has been {status}.')
    except Exception as e:
        logger.error(f"Error toggling user status: {str(e)}")
        messages.error(request, f'Error: {str(e)}')
    return redirect('dashboard')

# ============================================================================
# SESSION MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_superuser, login_url='/')
def dashboard_sessions(request):
    """User Sessions Dashboard - Superuser only"""
    try:
        sessions = UserSession.objects.select_related('user').all().order_by('-started_at')
        
        search_query = request.GET.get('search', '')
        if search_query:
            sessions = sessions.filter(
                Q(session_id__icontains=search_query) |
                Q(visitor_id__icontains=search_query) |
                Q(user__email__icontains=search_query) |
                Q(ip_address__icontains=search_query)
            )
        
        filter_type = request.GET.get('filter', '')
        if filter_type == 'authenticated':
            sessions = sessions.filter(is_authenticated=True)
        elif filter_type == 'anonymous':
            sessions = sessions.filter(is_authenticated=False)
        elif filter_type == 'bot':
            sessions = sessions.filter(is_bot=True)
        
        date_from = request.GET.get('date_from', '')
        if date_from:
            try:
                sessions = sessions.filter(started_at__gte=parse_date(date_from))
            except:
                pass
        
        sort_by = request.GET.get('sort', '-started_at')
        if sort_by in ['started_at', '-started_at']:
            sessions = sessions.order_by(sort_by)
        
        paginator = Paginator(sessions, 25)
        page_number = request.GET.get('page', 1)
        try:
            page_obj = paginator.page(page_number)
        except (PageNotAnInteger, EmptyPage):
            page_obj = paginator.page(1)
        
        try:
            total_sessions = UserSession.objects.count()
            authenticated_sessions = UserSession.objects.filter(is_authenticated=True).count()
            anonymous_sessions = UserSession.objects.filter(is_authenticated=False).count()
            bot_sessions = UserSession.objects.filter(is_bot=True).count()
            
            device_stats = UserSession.objects.values('device_type').annotate(
                count=Count('id')
            ).order_by('-count')
            
            top_countries = UserSession.objects.exclude(country='').values('country').annotate(
                count=Count('id')
            ).order_by('-count')[:10]
        except Exception as e:
            logger.error(f"Error getting session stats: {str(e)}")
            total_sessions = authenticated_sessions = anonymous_sessions = bot_sessions = 0
            device_stats = []
            top_countries = []
        
        context = {
            'page_title': 'Admin Dashboard - Sessions',
            'sessions': page_obj,
            'total_sessions': total_sessions,
            'authenticated_sessions': authenticated_sessions,
            'anonymous_sessions': anonymous_sessions,
            'bot_sessions': bot_sessions,
            'device_stats': device_stats,
            'top_countries': top_countries,
            'search_query': search_query,
            'filter_type': filter_type,
            'date_from': date_from,
        }
        
        return render(request, 'website/dashboard_sessions.html', context)
    except Exception as e:
        logger.error(f"Error in dashboard_sessions: {str(e)}")
        messages.error(request, f'Error loading sessions: {str(e)}')
        return render(request, 'website/dashboard_sessions.html', {
            'page_title': 'Admin Dashboard - Sessions',
            'error': str(e),
            'sessions': [],
        })

@login_required
@user_passes_test(is_superuser, login_url='/')
def session_detail(request, session_id):
    """Detailed view of a user session"""
    try:
        session = get_object_or_404(UserSession.objects.select_related('user'), id=session_id)
        
        try:
            activities = session.activities.all().order_by('-timestamp')[:100]
            activity_stats = activities.values('event_type').annotate(count=Count('id')).order_by('-count')
            
            response_times = [a.response_time_ms for a in activities if a.response_time_ms]
            avg_response_time = sum(response_times) / len(response_times) if response_times else 0
            total_activities = session.activities.count()
        except Exception as e:
            logger.error(f"Error getting session activities: {str(e)}")
            activities = []
            activity_stats = []
            avg_response_time = 0
            total_activities = 0
        
        context = {
            'page_title': f'Session Details',
            'session': session,
            'activities': activities,
            'activity_stats': activity_stats,
            'avg_response_time': avg_response_time,
            'total_activities': total_activities,
        }
        
        return render(request, 'website/session_detail.html', context)
    except Exception as e:
        logger.error(f"Error in session_detail: {str(e)}")
        messages.error(request, f'Error: {str(e)}')
        return redirect('dashboard_sessions')

@login_required
@user_passes_test(is_superuser, login_url='/')
@require_http_methods(["POST"])
def delete_session(request, session_id):
    """Delete a session"""
    try:
        session = get_object_or_404(UserSession, id=session_id)
        session_id_str = session.session_id[:20]
        session.delete()
        messages.success(request, f'Session {session_id_str}... deleted.')
    except Exception as e:
        logger.error(f"Error deleting session: {str(e)}")
        messages.error(request, f'Error: {str(e)}')
    return redirect('dashboard_sessions')

# ============================================================================
# ACTIVITY MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(is_superuser, login_url='/')
def dashboard_activities(request):
    """User Activities Dashboard - Superuser only"""
    try:
        activities = UserActivity.objects.select_related('session', 'session__user').all().order_by('-timestamp')
        
        search_query = request.GET.get('search', '')
        if search_query:
            activities = activities.filter(
                Q(path__icontains=search_query) |
                Q(url__icontains=search_query) |
                Q(session__user__email__icontains=search_query)
            )
        
        filter_type = request.GET.get('filter', '')
        if filter_type in ['page_view', 'api_request', 'interaction', 'auth', 'custom_event']:
            activities = activities.filter(event_type=filter_type)
        
        status_code = request.GET.get('status_code', '')
        if status_code:
            try:
                activities = activities.filter(status_code=int(status_code))
            except:
                pass
        
        date_from = request.GET.get('date_from', '')
        if date_from:
            try:
                activities = activities.filter(timestamp__gte=parse_date(date_from))
            except:
                pass
        
        sort_by = request.GET.get('sort', '-timestamp')
        if sort_by in ['timestamp', '-timestamp']:
            activities = activities.order_by(sort_by)
        
        paginator = Paginator(activities, 50)
        page_number = request.GET.get('page', 1)
        try:
            page_obj = paginator.page(page_number)
        except (PageNotAnInteger, EmptyPage):
            page_obj = paginator.page(1)
        
        try:
            total_activities = UserActivity.objects.count()
            page_views = UserActivity.objects.filter(event_type='page_view').count()
            api_requests = UserActivity.objects.filter(event_type='api_request').count()
            interactions = UserActivity.objects.filter(event_type='interaction').count()
            auth_events = UserActivity.objects.filter(event_type='auth').count()
            
            event_type_stats = list(UserActivity.objects.values('event_type').annotate(
                count=Count('id')
            ).order_by('-count'))
            # Add display names for event types
            for stat in event_type_stats:
                stat['event_type_display'] = stat['event_type'].replace('_', ' ').title()
            
            status_code_stats = UserActivity.objects.exclude(
                status_code__isnull=True
            ).values('status_code').annotate(count=Count('id')).order_by('-status_code')[:20]
            
            response_times = UserActivity.objects.exclude(response_time_ms__isnull=True)
            avg_response_time = response_times.aggregate(Avg('response_time_ms'))['response_time_ms__avg'] or 0
        except Exception as e:
            logger.error(f"Error getting activity stats: {str(e)}")
            total_activities = page_views = api_requests = interactions = auth_events = 0
            event_type_stats = []
            status_code_stats = []
            avg_response_time = 0
        
        context = {
            'page_title': 'Admin Dashboard - Activities',
            'activities': page_obj,
            'total_activities': total_activities,
            'page_views': page_views,
            'api_requests': api_requests,
            'interactions': interactions,
            'auth_events': auth_events,
            'event_type_stats': event_type_stats,
            'status_code_stats': status_code_stats,
            'avg_response_time': avg_response_time,
            'search_query': search_query,
            'filter_type': filter_type,
            'status_code': status_code,
            'sort_by': sort_by,
            'date_from': date_from,
        }
        
        return render(request, 'website/dashboard_activities.html', context)
    except Exception as e:
        logger.error(f"Error in dashboard_activities: {str(e)}")
        messages.error(request, f'Error loading activities: {str(e)}')
        return render(request, 'website/dashboard_activities.html', {
            'page_title': 'Admin Dashboard - Activities',
            'error': str(e),
            'activities': [],
        })

@login_required
@user_passes_test(is_superuser, login_url='/')
def activity_detail(request, activity_id):
    """Detailed view of a user activity"""
    try:
        activity = get_object_or_404(UserActivity.objects.select_related('session', 'session__user'), id=activity_id)
        
        query_params_json = json.dumps(activity.query_params, indent=2) if activity.query_params else None
        payload_json = json.dumps(activity.payload, indent=2) if activity.payload else None
        metadata_json = json.dumps(activity.metadata, indent=2) if activity.metadata else None
        
        context = {
            'page_title': f'Activity Details',
            'activity': activity,
            'query_params_json': query_params_json,
            'payload_json': payload_json,
            'metadata_json': metadata_json,
        }
        
        return render(request, 'website/activity_detail.html', context)
    except Exception as e:
        logger.error(f"Error in activity_detail: {str(e)}")
        messages.error(request, f'Error: {str(e)}')
        return redirect('dashboard_activities')

@login_required
@user_passes_test(is_superuser, login_url='/')
@require_http_methods(["POST"])
def delete_activity(request, activity_id):
    """Delete an activity"""
    try:
        activity = get_object_or_404(UserActivity, id=activity_id)
        activity_path = activity.path[:50]
        activity.delete()
        messages.success(request, f'Activity deleted.')
    except Exception as e:
        logger.error(f"Error deleting activity: {str(e)}")
        messages.error(request, f'Error: {str(e)}')
    return redirect('dashboard_activities')

# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

@login_required
@user_passes_test(is_superuser, login_url='/')
def export_users_excel(request):
    """Export users to Excel"""
    try:
        users = User.objects.all().order_by('-date_joined')
        
        search_query = request.GET.get('search', '')
        if search_query:
            users = users.filter(
                Q(email__icontains=search_query) |
                Q(username__icontains=search_query)
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        
        headers = ['ID', 'Email', 'Username', 'Verified', 'Active', 'Date Joined']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        for row_num, user in enumerate(users[:10000], 2):
            ws.cell(row=row_num, column=1, value=user.id)
            ws.cell(row=row_num, column=2, value=user.email)
            ws.cell(row=row_num, column=3, value=user.username)
            ws.cell(row=row_num, column=4, value='Yes' if user.is_verified else 'No')
            ws.cell(row=row_num, column=5, value='Yes' if user.is_active else 'No')
            ws.cell(row=row_num, column=6, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S'))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    except Exception as e:
        logger.error(f"Error exporting users: {str(e)}")
        messages.error(request, f'Error exporting: {str(e)}')
        return redirect('dashboard')

@login_required
@user_passes_test(is_superuser, login_url='/')
def export_sessions_excel(request):
    """Export sessions to Excel"""
    try:
        sessions = UserSession.objects.select_related('user').all().order_by('-started_at')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sessions"
        
        headers = ['ID', 'Session ID', 'User Email', 'Authenticated', 'IP Address', 'Country', 
                   'Device Type', 'Started At', 'Is Bot']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        for row_num, session in enumerate(sessions[:10000], 2):
            ws.cell(row=row_num, column=1, value=session.id)
            ws.cell(row=row_num, column=2, value=session.session_id[:30])
            ws.cell(row=row_num, column=3, value=session.user.email if session.user else 'Anonymous')
            ws.cell(row=row_num, column=4, value='Yes' if session.is_authenticated else 'No')
            ws.cell(row=row_num, column=5, value=str(session.ip_address) if session.ip_address else '')
            ws.cell(row=row_num, column=6, value=session.country or '')
            ws.cell(row=row_num, column=7, value=session.get_device_type_display())
            ws.cell(row=row_num, column=8, value=session.started_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=9, value='Yes' if session.is_bot else 'No')
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'sessions_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    except Exception as e:
        logger.error(f"Error exporting sessions: {str(e)}")
        messages.error(request, f'Error exporting: {str(e)}')
        return redirect('dashboard_sessions')

@login_required
@user_passes_test(is_superuser, login_url='/')
def export_activities_excel(request):
    """Export activities to Excel"""
    try:
        activities = UserActivity.objects.select_related('session', 'session__user').all().order_by('-timestamp')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activities"
        
        headers = ['ID', 'Event Type', 'Path', 'Method', 'Status Code', 'Response Time (ms)', 
                   'User Email', 'Client IP', 'Timestamp']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        for row_num, activity in enumerate(activities[:10000], 2):
            ws.cell(row=row_num, column=1, value=activity.id)
            ws.cell(row=row_num, column=2, value=activity.get_event_type_display())
            ws.cell(row=row_num, column=3, value=activity.path[:100] if activity.path else '')
            ws.cell(row=row_num, column=4, value=activity.method or '')
            ws.cell(row=row_num, column=5, value=activity.status_code or '')
            ws.cell(row=row_num, column=6, value=activity.response_time_ms or '')
            ws.cell(row=row_num, column=7, value=activity.session.user.email if activity.session and activity.session.user else 'Anonymous')
            ws.cell(row=row_num, column=8, value=str(activity.client_ip) if activity.client_ip else '')
            ws.cell(row=row_num, column=9, value=activity.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'activities_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    except Exception as e:
        logger.error(f"Error exporting activities: {str(e)}")
        messages.error(request, f'Error exporting: {str(e)}')
        return redirect('dashboard_activities')
